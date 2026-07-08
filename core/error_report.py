"""
error_report.py

Genera un reporte de errores en formato Excel.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from core.validator import ErrorValidacion


class ErrorReport:
    """
    Genera un archivo Excel con los errores encontrados
    durante la validación.
    """

    def generar(
        self,
        errores: list[ErrorValidacion],
        archivo_salida: str | Path,
    ):

        datos = []

        for error in errores:

            datos.append({

                "Fila Excel": error.fila_excel,

                "Día": error.dia,

                "Cliente": error.cliente,

                "Campo": error.campo,

                "Valor encontrado": error.valor,

                "Descripción": error.mensaje,

            })

        df = pd.DataFrame(datos)

        ruta = Path(archivo_salida)

        ruta.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        with pd.ExcelWriter(
            ruta,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="Errores",
                index=False,
            )

            hoja = writer.sheets["Errores"]

            self._formatear_hoja(hoja)

    def _formatear_hoja(self, hoja):

        color_encabezado = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        fuente = Font(

            bold=True,

            color="FFFFFF",

        )

        for celda in hoja[1]:

            celda.fill = color_encabezado

            celda.font = fuente

            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        hoja.freeze_panes = "A2"

        hoja.auto_filter.ref = hoja.dimensions

        for columna in hoja.columns:

            longitud = 0

            letra = get_column_letter(
                columna[0].column
            )

            for celda in columna:

                if celda.value:

                    longitud = max(
                        longitud,
                        len(str(celda.value))
                    )

            hoja.column_dimensions[
                letra
            ].width = min(
                longitud + 3,
                60
            )